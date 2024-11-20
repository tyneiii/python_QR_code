from openpyxl import load_workbook  # Import phần mềm openpyxl để làm việc với file Excel
from openpyxl.utils import get_column_letter  # Import hàm để chuyển đổi số cột sang chữ cái trong Excel
from datetime import datetime  # Import thư viện datetime để làm việc với thời gian và ngày tháng
from pyzbar.pyzbar import decode  # Import hàm decode từ thư viện pyzbar để giải mã mã QR
import cv2  # Import thư viện OpenCV để xử lý ảnh và video

def find_empty_column(sheet):
    # Hàm này tìm cột trống trong file Excel để ghi dữ liệu điểm danh
    # Dùng trong việc quét các cột và xác định cột trống
    max_col = sheet.max_column
    for col_num in range(3, max_col + 2):
        is_merged = False
        for range_ in sheet.merged_cells.ranges:
            if range_.min_col <= col_num <= range_.max_col:
                is_merged = True
                break
        if not any(sheet.cell(row=1, column=col_num).value for _ in range(sheet.max_row)) and not is_merged:
            return col_num
    return max_col + 1

def markAttendance(excel_file, code):
    # Hàm này thêm dữ liệu điểm danh vào file Excel dựa trên mã và ngày
    # Đánh dấu 'X' cho mã sinh viên nếu chưa được điểm danh trong ngày
    try:
        now = datetime.now()
        date_string = now.strftime('%Y-%m-%d')

        # Load file Excel
        workbook = load_workbook(excel_file)
        sheet = workbook.active

        # Kiểm tra xem ngày hôm nay đã được ghi chưa
        date_column = None
        for col_num in range(3, sheet.max_column + 1):
            date_cell = sheet.cell(row=1, column=col_num)
            if date_cell.value == date_string:
                date_column = col_num
                break

        if date_column is None:
            # Nếu ngày hôm nay chưa có, tìm cột trống để ghi 'X' vào
            empty_col = find_empty_column(sheet)

            # Ghi ngày tháng năm vào hàng đầu tiên của cột mới
            sheet.cell(row=1, column=empty_col, value=date_string)
            date_column = empty_col

        # Biến cờ để đảm bảo chỉ ghi 'X' nếu mã sinh viên chưa được đánh dấu trong ngày
        attended = False

        for row_num in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row_num, column=1)
            if cell.value == code:
                if sheet.cell(row=row_num, column=date_column).value != 'X':
                    sheet.cell(row=row_num, column=date_column, value='X')
                    attended = True
                    print(f"Attendance marked 'X' for Mã Sinh Viên: {code} on {date_string}")
                else:
                    print(f"Mã Sinh Viên: {code} has already been marked on {date_string}")
                break

        # Lưu lại file Excel
        workbook.save(excel_file)

    except Exception as e:
        print("Error marking attendance:", e)


def decode_img(img, excel_file):
    # Hàm này giải mã ảnh từ camera và gọi hàm markAttendance để điểm danh
    # Chuyển đổi mã QR thành mã số nguyên và gửi đi để điểm danh
    try:
        gray_img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        barcodes = decode(gray_img)

        for barcode in barcodes:
            myData = barcode.data.decode("utf-8")
            markAttendance(excel_file, int(myData))  # Chuyển đổi sang số nguyên từ QR code

    except Exception as e:
        print("Error decoding image:", e)
